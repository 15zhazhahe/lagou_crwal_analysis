"""
将excel格式的文件，存入至mysql数据库中
"""
import os
import pymysql
from openpyxl import load_workbook
# 获取excel路径
d = os.path.dirname(__file__)
parent_path = os.path.dirname(d)
xlsx_path = os.path.join(parent_path, 'result/拉勾网_数据挖掘.xlsx')

wb = load_workbook(xlsx_path)
ws = wb.get_sheet_by_name('Sheet')
max_row = ws.max_row
max_col = ws.max_column
print(max_row, max_col)

# 创建连接(host,user,keyword,database,charset)
db = pymysql.connect('localhost', 'root', '******', 'lagou_crawl', charset='utf8')
# 创建游标
cursor = db.cursor()
# 插入数据的sql语句
insert_sql = """
    INSERT INTO lagou_crawl.posiition_info(
        positionId, positionName, education, salary,
        workyear, firstType, secondType, positionLables,
        companyLabelList, companyId, companyShortName,
        companyFullName, industryField, companySize,
        financeStage, city, latitude, longitude) 
        VALUES ('%s', '%s', '%s', '%s',
                '%s', '%s', '%s', '%s',
                '%s', '%s', '%s', '%s',
                '%s', '%s', '%s', '%s',
                '%s', '%s')
"""
# 遍历excel,并将其存入mysql中
for i in range(2,max_row+1):
    data = []
    for j in range(1,max_col+1):
        data.append(ws.cell(row=i, column=j).value)

    insert = (insert_sql % (data[0], data[1], data[2], data[3],
                            data[4], data[5], data[6], data[7],
                            data[8], data[9], data[10], data[11],
                            data[12], data[13], data[14], data[15],
                            data[16], data[17]))
    print(insert)                        
    # 执行
    cursor.execute(insert)
    # 提交
    db.commit()
# 关闭
cursor.close()
db.close()